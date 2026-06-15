"""
XLSXImporter - Importador para arquivos Excel (.xlsx) e CSV.
==============================================================
Usa openpyxl para Excel e csv/pandas para CSV.
Concatena texto de células.
"""

from __future__ import annotations

import csv
import re
import unicodedata
from pathlib import Path
from typing import List, Sequence, Set

from .base_importer import BaseImporter, ImportResult, ImporterError
from ..utils.logger import get_logger


class XLSXImporter(BaseImporter):
    """
    Importador para arquivos Excel (.xlsx) e CSV.
    
    Características:
    - Suporta múltiplas planilhas
    - Extrai texto de todas as células
    - Suporta CSV com detecção de delimitador
    """
    
    SUPPORTED_EXTENSIONS: List[str] = ['.xlsx', '.xls', '.csv', '.tsv']
    TEXT_COLUMN_HINTS = {
        'texto', 'text', 'conteudo', 'content', 'corpus',
        'documento', 'document', 'mensagem', 'message',
        'resposta', 'fala', 'comentario', 'comentarios',
    }
    ID_COLUMN_HINTS = {
        'id', 'codigo', 'cod', 'code', 'uuid', 'pk', 'key',
        'registro', 'indice', 'index', 'protocolo', 'matricula',
    }
    
    def __init__(self) -> None:
        super().__init__()
        self._logger = get_logger(__name__)
        self._openpyxl = None
    
    def _ensure_openpyxl(self) -> None:
        """Importa openpyxl sob demanda."""
        if self._openpyxl is None:
            try:
                import openpyxl
                self._openpyxl = openpyxl
            except ImportError:
                raise ImporterError(
                    what="Biblioteca openpyxl não encontrada.",
                    why="O openpyxl é necessário para ler arquivos Excel.",
                    how="Instale com: pip install openpyxl"
                )
    
    def can_handle(self, file_path: str) -> bool:
        """
        Verifica se o arquivo é Excel ou CSV.
        
        Args:
            file_path: Caminho do arquivo.
            
        Returns:
            True se for um arquivo suportado.
        """
        ext = self._get_file_extension(file_path)
        return ext in self.SUPPORTED_EXTENSIONS
    
    def extract(self, file_path: str) -> ImportResult:
        """
        Extrai texto de um arquivo Excel ou CSV.
        
        Args:
            file_path: Caminho do arquivo.
            
        Returns:
            ImportResult com o texto e metadados.
            
        Raises:
            ImporterError: Se houver erro na extração.
        """
        path = self._validate_file_exists(file_path)
        ext = self._get_file_extension(file_path)
        
        if ext in ['.csv', '.tsv']:
            return self._extract_csv(path, ext)
        elif ext == '.xls':
            raise ImporterError(
                what="Formato .xls (Excel 97-2003) não é suportado diretamente.",
                why="O formato antigo do Excel (.xls) requer bibliotecas adicionais.",
                how="Abra o arquivo no Excel e salve como .xlsx (formato moderno)."
            )
        else:
            return self._extract_xlsx(path)
    
    def _extract_xlsx(self, path: Path) -> ImportResult:
        """
        Extrai texto de arquivo Excel (.xlsx).
        
        Args:
            path: Path do arquivo.
            
        Returns:
            ImportResult com o texto.
        """
        self._ensure_openpyxl()
        warnings: List[str] = []
        
        try:
            workbook = self._openpyxl.load_workbook(path, data_only=True)
        except Exception as e:
            error_msg = str(e).lower()
            
            if "bad zip file" in error_msg or "not a valid" in error_msg:
                raise ImporterError(
                    what="O arquivo não parece ser uma planilha Excel válida.",
                    why="O conteúdo não corresponde ao formato .xlsx esperado.",
                    how="Verifique se o arquivo é realmente uma planilha Excel."
                )
            
            self._logger.error(f"Erro ao abrir Excel: {e}")
            raise ImporterError(
                what="Erro ao abrir a planilha Excel.",
                why=str(e),
                how="Verifique se o arquivo é um .xlsx válido e não está corrompido."
            )
        
        all_text: List[str] = []
        all_iramuteq_docs: List[str] = []
        sheet_count = 0
        total_cells = 0
        total_rows = 0
        total_iramuteq_docs = 0
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            sheet_count += 1
            table_rows: List[List[str]] = []
            
            for row in sheet.iter_rows(values_only=True):
                row_values = [self._cell_to_string(value) for value in row]
                if any(row_values):
                    table_rows.append(row_values)
                    total_cells += sum(1 for value in row_values if value)
            
            if table_rows:
                total_rows += len(table_rows)
                sheet_text = self._table_to_plain_text(table_rows)
                all_text.append(f"[Planilha: {sheet_name}]\n{sheet_text}")
                sheet_iramuteq = self._table_to_iramuteq_text(
                    table_rows,
                    source_label=sheet_name,
                )
                if sheet_iramuteq:
                    all_iramuteq_docs.append(sheet_iramuteq)
                    total_iramuteq_docs += sheet_iramuteq.count("**** ")
        
        workbook.close()
        
        text = "\n\n".join(all_text)
        iramuteq_text = "\n\n".join(all_iramuteq_docs)
        
        if not text.strip():
            warnings.append("A planilha está vazia ou não contém texto.")
        if not iramuteq_text.strip():
            warnings.append(
                "A planilha não gerou documentos IRaMuTeQ válidos. "
                "Verifique se as linhas possuem conteúdo textual."
            )
        
        if sheet_count > 1:
            warnings.append(
                f"Arquivo contém {sheet_count} planilhas. "
                "O texto de todas foi concatenado."
            )
        
        words = text.split()
        metadata = {
            'sheet_count': sheet_count,
            'row_count': total_rows,
            'cell_count': total_cells,
            'word_count': len(words),
            'char_count': len(text),
            'iramuteq_doc_count': total_iramuteq_docs,
            'iramuteq_text': iramuteq_text,
            'file_size': self._get_file_size(str(path)),
            'file_size_formatted': self._format_file_size(self._get_file_size(str(path))),
        }
        
        self._logger.info(
            f"Excel importado: {metadata['word_count']} palavras, "
            f"{sheet_count} planilhas, {total_cells} células"
        )
        
        return ImportResult(
            text=text,
            source_file=str(path),
            encoding="utf-8",
            warnings=warnings,
            metadata=metadata,
        )
    
    def _extract_csv(self, path: Path, ext: str) -> ImportResult:
        """
        Extrai texto de arquivo CSV/TSV.
        
        Args:
            path: Path do arquivo.
            ext: Extensão do arquivo.
            
        Returns:
            ImportResult com o texto.
        """
        warnings: List[str] = []
        
        # Detectar encoding
        encoding = self.detect_encoding(str(path))
        
        # Determinar delimitador
        if ext == '.tsv':
            delimiter = '\t'
        else:
            delimiter = self._detect_delimiter(path, encoding)
        
        try:
            with open(path, 'r', encoding=encoding, newline='') as f:
                reader = csv.reader(f, delimiter=delimiter)
                table_rows: List[List[str]] = []
                total_cells = 0
                
                for row in reader:
                    row_values = [cell.strip() for cell in row]
                    if any(row_values):
                        table_rows.append(row_values)
                        total_cells += sum(1 for value in row_values if value)
        
        except UnicodeDecodeError:
            # Tentar com latin-1
            with open(path, 'r', encoding='latin-1', newline='') as f:
                reader = csv.reader(f, delimiter=delimiter)
                table_rows = []
                total_cells = 0
                
                for row in reader:
                    row_values = [cell.strip() for cell in row]
                    if any(row_values):
                        table_rows.append(row_values)
                        total_cells += sum(1 for value in row_values if value)
            
            encoding = 'latin-1'
            warnings.append(f"Encoding alternativo utilizado: {encoding}")
        
        except Exception as e:
            self._logger.error(f"Erro ao ler CSV: {e}")
            raise ImporterError(
                what="Erro ao ler o arquivo CSV.",
                why=str(e),
                how="Verifique se o arquivo é um CSV válido com formatação correta."
            )
        
        text = self._table_to_plain_text(table_rows)
        iramuteq_text = self._table_to_iramuteq_text(
            table_rows,
            source_label=path.stem,
        )
        
        if not text.strip():
            warnings.append("O arquivo CSV está vazio ou não contém texto.")
        if not iramuteq_text.strip():
            warnings.append(
                "O CSV não gerou documentos IRaMuTeQ válidos. "
                "Verifique se as linhas possuem conteúdo textual."
            )
        
        words = text.split()
        metadata = {
            'row_count': len(table_rows),
            'cell_count': total_cells,
            'delimiter': repr(delimiter),
            'word_count': len(words),
            'char_count': len(text),
            'iramuteq_doc_count': iramuteq_text.count("**** "),
            'iramuteq_text': iramuteq_text,
            'file_size': self._get_file_size(str(path)),
            'file_size_formatted': self._format_file_size(self._get_file_size(str(path))),
        }
        
        self._logger.info(
            f"CSV importado: {metadata['word_count']} palavras, "
            f"{len(table_rows)} linhas"
        )
        
        return ImportResult(
            text=text,
            source_file=str(path),
            encoding=encoding,
            warnings=warnings,
            metadata=metadata,
        )

    def _cell_to_string(self, value) -> str:
        """Converte valor de celula para string limpa."""
        if value is None:
            return ""
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value).strip()

    def _table_to_plain_text(self, rows: Sequence[Sequence[str]]) -> str:
        """Converte tabela em texto simples, uma linha por registro."""
        plain_rows: List[str] = []
        for row in rows:
            row_values = [value for value in row if value]
            if row_values:
                plain_rows.append(" ".join(row_values))
        return "\n".join(plain_rows)

    def _table_to_iramuteq_text(
        self,
        rows: Sequence[Sequence[str]],
        source_label: str,
    ) -> str:
        """
        Converte linhas tabulares em corpus IRaMuTeQ.

        Cada linha de dados vira um documento:
        **** *doc_1 *fonte_sheet *var_valor
        Texto da linha.
        """
        if not rows:
            return ""

        next_row = rows[1] if len(rows) > 1 else None
        header_is_present = self._looks_like_header(rows[0], next_row)
        header = list(rows[0]) if header_is_present else []
        data_rows = rows[1:] if header_is_present else rows

        max_cols = max(len(row) for row in rows)
        if not header:
            header = [f"col_{idx + 1}" for idx in range(max_cols)]
        elif len(header) < max_cols:
            header = header + [f"col_{idx + 1}" for idx in range(len(header), max_cols)]

        normalized_header = [
            self._slugify(name, default=f"col_{idx + 1}", max_len=24)
            for idx, name in enumerate(header)
        ]
        text_columns = {
            idx
            for idx, name in enumerate(normalized_header)
            if name in self.TEXT_COLUMN_HINTS
        }
        variable_columns = self._infer_variable_columns(
            data_rows=data_rows,
            normalized_header=normalized_header,
            text_columns=text_columns,
        )

        source_token = self._slugify(source_label, default="planilha", max_len=24)
        documents: List[str] = []

        for row_index, row in enumerate(data_rows, start=1):
            aligned_row = list(row) + [""] * (len(normalized_header) - len(row))
            aligned_row = aligned_row[:len(normalized_header)]

            text_parts = [aligned_row[idx] for idx in text_columns if aligned_row[idx]]
            if not text_parts:
                text_parts = [value for value in aligned_row if value]

            body = " ".join(text_parts).strip()
            if not body:
                continue

            variables: List[str] = [f"*doc_{row_index}", f"*fonte_{source_token}"]
            for idx, value in enumerate(aligned_row):
                if not value or idx in text_columns or idx not in variable_columns:
                    continue
                value_token = self._slugify(value, default="", max_len=36)
                if not value_token:
                    continue
                variables.append(f"*{normalized_header[idx]}_{value_token}")

            variables = list(dict.fromkeys(variables))
            command_line = f"**** {' '.join(variables)}"
            documents.append(f"{command_line}\n{body}")

        return "\n\n".join(documents)

    def _infer_variable_columns(
        self,
        data_rows: Sequence[Sequence[str]],
        normalized_header: Sequence[str],
        text_columns: Set[int],
    ) -> Set[int]:
        """
        Seleciona colunas candidatas a variáveis semânticas.

        Evita colunas identificadoras de alta cardinalidade (ex.: id único).
        """
        if not data_rows:
            return set()

        selected: Set[int] = set()
        n_rows = len(data_rows)
        for col_idx, header_name in enumerate(normalized_header):
            if col_idx in text_columns:
                continue

            values: List[str] = []
            for row in data_rows:
                value = row[col_idx].strip() if col_idx < len(row) else ""
                if value:
                    values.append(value)

            if not values:
                continue

            unique_count = len(set(values))
            cardinality_ratio = unique_count / max(1, len(values))
            numeric_ratio = sum(1 for value in values if value.isdigit()) / max(1, len(values))

            is_id_like = (
                header_name in self.ID_COLUMN_HINTS
                or header_name.endswith("_id")
                or header_name.startswith("id_")
            )
            high_cardinality = len(values) >= 10 and cardinality_ratio >= 0.90
            mostly_numeric = numeric_ratio >= 0.80

            if is_id_like and (high_cardinality or mostly_numeric):
                continue

            if high_cardinality and mostly_numeric:
                continue

            selected.add(col_idx)

        return selected

    def _looks_like_header(
        self,
        row: Sequence[str],
        next_row: Sequence[str] | None = None,
    ) -> bool:
        """Heuristica simples para detectar cabecalho de colunas."""
        values = [value.strip() for value in row if value and value.strip()]
        if len(values) < 2:
            return False

        normalized_values = [
            self._slugify(value, default="", max_len=40)
            for value in values
        ]
        if any(value in self.TEXT_COLUMN_HINTS for value in normalized_values):
            return True

        unique_values = len(set(value.lower() for value in values))
        if unique_values != len(values):
            return False

        # Sem segunda linha de referencia, evitar falso-positivo.
        if next_row is None:
            return False

        next_values = [value.strip() for value in next_row if value and value.strip()]
        if not next_values:
            return False

        short_label_like = sum(
            1
            for value in values
            if len(value) <= 30
            and ' ' not in value
            and not any(char.isdigit() for char in value)
        )
        data_like_next = sum(
            1
            for value in next_values
            if (' ' in value) or any(char.isdigit() for char in value) or len(value) > 30
        )
        return (
            short_label_like / len(values) >= 0.6
            and data_like_next / len(next_values) >= 0.3
        )

    def _slugify(self, value: str, default: str, max_len: int = 32) -> str:
        """Normaliza string para token IRaMuTeQ sem acentos e simbolos."""
        normalized = unicodedata.normalize('NFD', value or "")
        normalized = ''.join(
            char for char in normalized
            if unicodedata.category(char) != 'Mn'
        )
        normalized = normalized.lower()
        normalized = re.sub(r'[^a-z0-9]+', '_', normalized)
        normalized = re.sub(r'_+', '_', normalized).strip('_')
        if max_len > 0:
            normalized = normalized[:max_len].strip('_')
        return normalized or default
    
    def _detect_delimiter(self, path: Path, encoding: str) -> str:
        """
        Detecta o delimitador de um arquivo CSV.
        
        Args:
            path: Path do arquivo.
            encoding: Encoding do arquivo.
            
        Returns:
            Delimitador detectado.
        """
        try:
            with open(path, 'r', encoding=encoding) as f:
                sample = f.read(8192)
            
            # Usar Sniffer do CSV
            try:
                dialect = csv.Sniffer().sniff(sample, delimiters=',;\t|')
                return dialect.delimiter
            except csv.Error:
                pass
            
            # Fallback: contar ocorrências
            delimiters = {',': 0, ';': 0, '\t': 0, '|': 0}
            for char in sample:
                if char in delimiters:
                    delimiters[char] += 1
            
            # Retornar o mais comum (se houver)
            most_common = max(delimiters, key=delimiters.get)
            if delimiters[most_common] > 0:
                return most_common
            
        except Exception:
            pass
        
        # Default: vírgula
        return ','
